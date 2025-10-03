"""
Excel parser for raw data ingestion using openpyxl.
This module provides a function to read Excel files (.xlsx) from either a local
file path or a URL, and return the data in memory as Python dictionaries.
"""

from openpyxl import load_workbook   # openpyxl loads Excel workbooks and sheets
import requests                      # requests lets us download files from URLs
from io import BytesIO               # BytesIO lets us treat downloaded bytes as a file


def read_excel(source: str, sheet_name: str = None):
    """
    Read Excel data from a local file path or URL.

    Args:
        source (str): Path to local file (e.g., "data.xlsx") or
                      HTTP(S) URL (e.g., "http://example.com/data.xlsx").
        sheet_name (str, optional): Specific sheet name to read.
                                    If None, the function will read *all* sheets.

    Returns:
        dict: A dictionary with this structure:
              {
                "SheetName1": [ {"col1": value1, "col2": value2, ...}, ... ],
                "SheetName2": [ {"col1": value1, "col2": value2, ...}, ... ],
              }
              - Each sheet becomes a key.
              - Each value is a list of row dictionaries.
              - The first row of the sheet is treated as headers (column names).

    Raises:
        ValueError: If the Excel file cannot be loaded (bad path, bad URL, corrupt file).
    """

    try:
        # CASE 1: Source is a URL (starts with http/https)
        if source.startswith("http"):
            # Download file into memory
            resp = requests.get(source, timeout=10)  # timeout avoids hanging
            resp.raise_for_status()                  # raise HTTPError if 404, 500, etc.
            # Load workbook from bytes as if it were a file on disk
            wb = load_workbook(filename=BytesIO(resp.content), data_only=True)
        else:
            # CASE 2: Source is a local file path
            wb = load_workbook(filename=source, data_only=True)

    except Exception as e:
        # Wrap *any* error (network, missing file, corrupt Excel) as ValueError
        raise ValueError(f"Failed to load Excel file: {e}")

    # Determine which sheets to read:
    # - If sheet_name was given, only read that one.
    # - Otherwise, read every sheet in the workbook.
    sheets_to_read = [sheet_name] if sheet_name else wb.sheetnames

    # Container for results
    data = {}

    # Loop through each requested sheet
    for sheet in sheets_to_read:
        try:
            ws = wb[sheet]   # ws = "worksheet" object
        except KeyError:
            raise ValueError(f"Unknown sheet: {sheet}")

        # First row in Excel (row 1) is assumed to be column headers
        headers = [cell.value for cell in ws[1]]

        # Storage for this sheet's rows
        rows = []

        # Start at row 2 (skip header row), read values only (no formatting)
        for row in ws.iter_rows(min_row=2, values_only=True):
            # Skip fully empty rows (rows containing only None values)
            if all(cell is None for cell in row):
                continue
            
            # Pair headers with row values → turn into dict
            # Example: headers = ["Name","Age"], row = ("Alice", 30)
            # Result: {"Name": "Alice", "Age": 30}
            row_dict = dict(zip(headers, row))
            rows.append(row_dict)

        # Save parsed rows under this sheet's name
        data[sheet] = rows

    # Return dict-of-lists-of-dicts
    return data